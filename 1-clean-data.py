# ----------------------------------------------------------------------------------------------------------------------
# LOAD REQUIRED LIBRARIES, DIRECTORIES, ETC. ---------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import datetime as dt

input_dir = r'D:\# Jvasco\Visualization\Vale de Cavalos'

import os
import sys
import openpyxl
import sqlalchemy
import traitlets_pcse
import requests
import xlrd
# del sys.path[0:20]
# path_to_the_model = os.path.abspath(os.path.join(os.getcwd(), '# WOFOST_v7.1/B) NEW_NPK_SEPT 2019'))
# sys.path.append(path_to_the_model)
import pcse
from pcse.util import ea_from_tdew
from pcse.db import NASAPowerWeatherDataProvider
from pcse.db import AgERA5WeatherDataProvider

# ----------------------------------------------------------------------------------------------------------------------
# Process data Vale de Cavalos -----------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------

# Load years -----------------------------------------------------------------------------------------------------------
TBASE = 10      # para vinha....
station_VC_2009 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2009.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2009['RAIN_CUM'] = np.cumsum(station_VC_2009.RR)
station_VC_2009['ETP_CUM'] = np.cumsum(station_VC_2009.ETP)
station_VC_2009['GDD'] = np.where(((station_VC_2009.TX + station_VC_2009.TN)/2) - TBASE < 0, 0, (station_VC_2009.TX + station_VC_2009.TN)/2 - TBASE)
station_VC_2009['GDD_CUM'] = np.cumsum(station_VC_2009.GDD)
station_VC_2010 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2010.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2010['RAIN_CUM'] = np.cumsum(station_VC_2010.RR)
station_VC_2010['ETP_CUM'] = np.cumsum(station_VC_2010.ETP)
station_VC_2010['GDD'] = np.where(((station_VC_2010.TX + station_VC_2010.TN)/2) - TBASE < 0, 0, (station_VC_2010.TX + station_VC_2010.TN)/2 - TBASE)
station_VC_2010['GDD_CUM'] = np.cumsum(station_VC_2010.GDD)
station_VC_2011 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2011.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2011['RAIN_CUM'] = np.cumsum(station_VC_2011.RR)
station_VC_2011['ETP_CUM'] = np.cumsum(station_VC_2011.ETP)
station_VC_2011['GDD'] = np.where(((station_VC_2011.TX + station_VC_2011.TN)/2) - TBASE < 0, 0, (station_VC_2011.TX + station_VC_2011.TN)/2 - TBASE)
station_VC_2011['GDD_CUM'] = np.cumsum(station_VC_2011.GDD)
station_VC_2012 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2012.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2012['RAIN_CUM'] = np.cumsum(station_VC_2012.RR)
station_VC_2012['ETP_CUM'] = np.cumsum(station_VC_2012.ETP)
station_VC_2012['GDD'] = np.where(((station_VC_2012.TX + station_VC_2012.TN)/2) - TBASE < 0, 0, (station_VC_2012.TX + station_VC_2012.TN)/2 - TBASE)
station_VC_2012['GDD_CUM'] = np.cumsum(station_VC_2012.GDD)
station_VC_2013 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2013_new.xlsx".format(input_dir), sheet_name='2013')
station_VC_2013['RAIN_CUM'] = np.cumsum(station_VC_2013.RR)
station_VC_2013['ETP_CUM'] = np.cumsum(station_VC_2013.ETP)
station_VC_2013['GDD'] = np.where(((station_VC_2013.TX + station_VC_2013.TN)/2) - TBASE < 0, 0, (station_VC_2013.TX + station_VC_2013.TN)/2 - TBASE)
station_VC_2013['GDD_CUM'] = np.cumsum(station_VC_2013.GDD)
station_VC_2014 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2014.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2014['RAIN_CUM'] = np.cumsum(station_VC_2014.RR)
station_VC_2014['ETP_CUM'] = np.cumsum(station_VC_2014.ETP)
station_VC_2014['GDD'] = np.where(((station_VC_2014.TX + station_VC_2014.TN)/2) - TBASE < 0, 0, (station_VC_2014.TX + station_VC_2014.TN)/2 - TBASE)
station_VC_2014['GDD_CUM'] = np.cumsum(station_VC_2014.GDD)
station_VC_2015 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2015.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2015['RAIN_CUM'] = np.cumsum(station_VC_2015.RR)
station_VC_2015['ETP_CUM'] = np.cumsum(station_VC_2015.ETP)
station_VC_2015['GDD'] = np.where(((station_VC_2015.TX + station_VC_2015.TN)/2) - TBASE < 0, 0, (station_VC_2015.TX + station_VC_2015.TN)/2 - TBASE)
station_VC_2015['GDD_CUM'] = np.cumsum(station_VC_2015.GDD)
station_VC_2016 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2016.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2016['RAIN_CUM'] = np.cumsum(station_VC_2016.RR)
station_VC_2016['ETP_CUM'] = np.cumsum(station_VC_2016.ETP)
station_VC_2016['GDD'] = np.where(((station_VC_2016.TX + station_VC_2016.TN)/2) - TBASE < 0, 0, (station_VC_2016.TX + station_VC_2016.TN)/2 - TBASE)
station_VC_2016['GDD_CUM'] = np.cumsum(station_VC_2016.GDD)
station_VC_2017 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2017.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2017['RAIN_CUM'] = np.cumsum(station_VC_2017.RR)
station_VC_2017['ETP_CUM'] = np.cumsum(station_VC_2017.ETP)
station_VC_2017['GDD'] = np.where(((station_VC_2017.TX + station_VC_2017.TN)/2) - TBASE < 0, 0, (station_VC_2017.TX + station_VC_2017.TN)/2 - TBASE)
station_VC_2017['GDD_CUM'] = np.cumsum(station_VC_2017.GDD)
station_VC_2018 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2018.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2018['RAIN_CUM'] = np.cumsum(station_VC_2018.RR)
station_VC_2018['ETP_CUM'] = np.cumsum(station_VC_2018.ETP)
station_VC_2018['GDD'] = np.where(((station_VC_2018.TX + station_VC_2018.TN)/2) - TBASE < 0, 0, (station_VC_2018.TX + station_VC_2018.TN)/2 - TBASE)
station_VC_2018['GDD_CUM'] = np.cumsum(station_VC_2018.GDD)
station_VC_2019 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2019.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2019['RAIN_CUM'] = np.cumsum(station_VC_2019.RR)
station_VC_2019['ETP_CUM'] = np.cumsum(station_VC_2019.ETP)
station_VC_2019['GDD'] = np.where(((station_VC_2019.TX + station_VC_2019.TN)/2) - TBASE < 0, 0, (station_VC_2019.TX + station_VC_2019.TN)/2 - TBASE)
station_VC_2019['GDD_CUM'] = np.cumsum(station_VC_2019.GDD)
station_VC_2020 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2020.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2020['RAIN_CUM'] = np.cumsum(station_VC_2020.RR)
station_VC_2020['ETP_CUM'] = np.cumsum(station_VC_2020.ETP)
station_VC_2020['GDD'] = np.where(((station_VC_2020.TX + station_VC_2020.TN)/2) - TBASE < 0, 0, (station_VC_2020.TX + station_VC_2020.TN)/2 - TBASE)
station_VC_2020['GDD_CUM'] = np.cumsum(station_VC_2020.GDD)
station_VC_2021 = pd.read_excel("{}/data/weather-data/weather-station/ETP_VC_2021.xlsx".format(input_dir), sheet_name='DataExp')
station_VC_2021['RAIN_CUM'] = np.cumsum(station_VC_2021.RR)
station_VC_2021['ETP_CUM'] = np.cumsum(station_VC_2021.ETP)
station_VC_2021['GDD'] = np.where(((station_VC_2021.TX + station_VC_2021.TN)/2) - TBASE < 0, 0, (station_VC_2021.TX + station_VC_2021.TN)/2 - TBASE)
station_VC_2021['GDD_CUM'] = np.cumsum(station_VC_2021.GDD)

# Process data ---------------------------------------------------------------------------------------------------------
station_VC = station_VC_2009.append(station_VC_2010).append(station_VC_2011).append(station_VC_2012).append(station_VC_2013).append(station_VC_2014).append(station_VC_2015).append(station_VC_2016).append(station_VC_2017).append(station_VC_2018).append(station_VC_2019).append(station_VC_2020).append(station_VC_2021)
station_VC['datetime'] = pd.to_datetime(station_VC['Date'], format='%d-%m-%Y', errors='coerce')
station_VC['Year'] = station_VC['datetime'].dt.year
station_VC['Country'] = 'Portugal'
station_VC['Station'] = 'Vale de Cavalos'
station_VC['CountryStation'] = station_VC['Country'].astype(str) + '_' + station_VC['Station'].astype(str)
station_VC['LAT'] = 39.28837262
station_VC['LONG'] = -8.5221237
station_VC['ELEV'] = 65
station_VC['Day'] = station_VC['datetime'].dt.day
station_VC['Month'] = station_VC['datetime'].dt.month
station_VC['M_month'] = 'M_' + station_VC['Month'].astype(str)
station_VC['Year'] = station_VC['datetime'].dt.year
station_VC['Y_year'] = 'Y_' + station_VC['Year'].astype(str)
station_VC['DOY'] = station_VC['datetime'].dt.dayofyear
station_VC['IRRAD_MJm2day'] = station_VC['RAYGL'] / 100000
station_VC['IRRAD_kJm2day'] = station_VC['IRRAD_MJm2day'] * 1000
station_VC['UM'] = np.where(station_VC['UM'] == '<<<<', 75, station_VC['UM'])
station_VC['UM'] = station_VC['UM'].astype(float)
station_VC['SNOWDEPTH_cm'] = -999
station_VC = station_VC.rename(columns={'TN': 'TMIN', 'TX': 'TMAX', 'RR': 'RAIN_mm', 'ETP': 'ET0_mm', 'VX': 'WIND_ms'})
station_VC['VAP_kPa'] = station_VC['TMIN'].apply(ea_from_tdew)
station_VC['Source'] = 'JSilva'
station_VC['Angstrom_A'] = -0.228
station_VC['Angstrom_B'] = -0.538
station_VC.to_csv('D:\# Jvasco\Visualization\Vale de Cavalos\data\weather-data\weather-station\Weather Data.csv')

# NASA Power -----------------------------------------------------------------------------------------------------------
# NASA_POWER = NASAPowerWeatherDataProvider(latitude=39.28837262, longitude=-8.5221237, force_update=True)
# station_VC['Angstrom_A'] = NASA_POWER.angstA
# station_VC['Angstrom_B'] = NASA_POWER.angstB

# AgERA5 ---------------------------------------------------------------------------------------------------------------
# weather_era5 = AgERA5WeatherDataProvider(latitude=39.28837262, longitude=-8.5221237, start_date=dt.date(2009, 1, 1), enddate=dt.date(2020, 12, 31))
# era5_export = pd.DataFrame(weather_era5.export()).set_index('DAY')
# era5_export.to_excel(os.path.join(input_dir, './Model Inputs/weather_agera5_vale de cavalos.xlsx'))

# Write WOFOST Weather Template ----------------------------------------------------------------------------------------
fn = r'D:\# Jvasco\Visualization\Vale de Cavalos\data\wofost-inputs\WOFOST_Weather_Vale de Cavalos.xlsx'
template = pd.read_excel(fn, header=None, sheet_name='ObservedWeather')
for station in station_VC.Station.unique():
    station_loop = station_VC[station_VC.Station == station]
    station_final = station_loop[['datetime', 'IRRAD_kJm2day', 'TMIN', 'TMAX', 'VAP_kPa', 'WIND_ms', 'RAIN_mm', 'SNOWDEPTH_cm']]
    sheet_name = station_loop['CountryStation'].unique()[0]
    writer = pd.ExcelWriter(fn, engine='openpyxl')
    book = load_workbook(fn)
    writer.book = book
    template.to_excel(writer, sheet_name=sheet_name, header=None, index=False)
    station_final.to_excel(writer, sheet_name=sheet_name, header=False, index=False, startcol=0, startrow=12)
    sheetname = book[sheet_name]
    sheetname.cell(row=2, column=2).value = station_loop['Country'].unique()[0]
    sheetname.cell(row=3, column=2).value = station_loop['Station'].unique()[0]
    sheetname.cell(row=4, column=2).value = 'Processed with WeatherData_Clean.py'
    sheetname.cell(row=5, column=2).value = station_loop['Source'].unique()[0]
    sheetname.cell(row=6, column=2).value = 'Joao Vasco Silva, WUR'
    sheetname.cell(row=9, column=1).value = station_loop['LONG'].unique()[0]
    sheetname.cell(row=9, column=2).value = station_loop['LAT'].unique()[0]
    sheetname.cell(row=9, column=3).value = station_loop['ELEV'].unique()[0]
    sheetname.cell(row=9, column=4).value = station_loop['Angstrom_A'].unique()[0]
    sheetname.cell(row=9, column=5).value = station_loop['Angstrom_B'].unique()[0]
    writer.save()

# ----------------------------------------------------------------------------------------------------------------------
# THE END --------------------------------------------------------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------------
